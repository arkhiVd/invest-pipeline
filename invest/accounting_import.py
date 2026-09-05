"""Strict Phase 11 Zerodha and Vested accounting importers."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import duckdb
from openpyxl import load_workbook

from invest import vested

ZERODHA_TRADE_HEADERS = (
    "symbol",
    "isin",
    "trade_date",
    "exchange",
    "segment",
    "series",
    "trade_type",
    "auction",
    "quantity",
    "price",
    "trade_id",
    "order_id",
    "order_execution_time",
)
ZERODHA_LEDGER_HEADERS = (
    "particulars",
    "posting_date",
    "cost_center",
    "voucher_type",
    "debit",
    "credit",
    "net_balance",
)
ZERODHA_DIVIDEND_HEADERS = (
    "Symbol",
    "Ex-date",
    "Qty",
    "Dividend per share",
    "Total dividend",
)
VESTED_SHEETS = {
    "All Transactions": (
        "Date",
        "Time (in UTC)",
        "Type",
        "Amount",
        "Account Balance",
        "Comment",
    ),
    "Trades": (
        "Date",
        "Time (in UTC)",
        "Name",
        "Ticker",
        "Activity",
        "Order Type",
        "Quantity",
        "Price Per Share (in USD)",
        "Cash Amount (in USD)",
        "Commission Charges (in USD)",
    ),
    "Transfers": ("Date", "Time (in UTC)", "Activity", "Cash Amount (in USD)"),
    "Income": (
        "Date",
        "Time (in UTC)",
        "Activity",
        "Ticker",
        "Gross Cash Amount (in USD)",
    ),
}


class AccountingImportError(RuntimeError):
    pass


def _decimal(value, label: str, *, positive: bool = False, nonnegative: bool = False):
    if isinstance(value, bool) or value is None or str(value).strip() == "":
        raise AccountingImportError(f"{label} must be numeric")
    try:
        result = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise AccountingImportError(f"{label} must be numeric") from exc
    if not result.is_finite():
        raise AccountingImportError(f"{label} must be finite")
    if positive and result <= 0:
        raise AccountingImportError(f"{label} must be positive")
    if nonnegative and result < 0:
        raise AccountingImportError(f"{label} must be nonnegative")
    return result


def _date(value, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise AccountingImportError(f"invalid {label}") from exc


def _event_time(day, value, label: str) -> datetime:
    day = _date(day, label)
    if isinstance(value, datetime):
        value = value.time()
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%I:%M:%S %p"):
        try:
            parsed = datetime.strptime(text, fmt).time()
            return datetime.combine(day, parsed, UTC)
        except ValueError:
            pass
    raise AccountingImportError(f"invalid {label} time")


def _hash(value) -> str:
    encoded = json.dumps(value, sort_keys=True, separators=(",", ":"), default=str).encode()
    return hashlib.sha256(encoded).hexdigest()


def _csv(path: Path, expected: tuple[str, ...]):
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8-sig")
        rows = list(csv.DictReader(text.splitlines()))
    except (UnicodeDecodeError, csv.Error) as exc:
        raise AccountingImportError("invalid CSV") from exc
    if tuple(rows[0].keys()) if rows else tuple() != expected:
        actual = tuple(rows[0].keys()) if rows else tuple()
        if actual != expected:
            raise AccountingImportError("CSV headers changed")
    if not rows:
        raise AccountingImportError("CSV has no data rows")
    return raw, rows


def parse_zerodha_tradebook(path: Path):
    raw, source = _csv(Path(path), ZERODHA_TRADE_HEADERS)
    rows = []
    seen = set()
    for row in source:
        symbol = row["symbol"].strip()
        isin = row["isin"].strip()
        exchange = row["exchange"].strip()
        segment = row["segment"].strip()
        if not symbol or not exchange or segment not in {"EQ", "MF"}:
            raise AccountingImportError("invalid trade identity")
        identity = isin or f"SYMBOL:{exchange}:{symbol}"
        if row["trade_type"] not in {"buy", "sell"} or row["auction"] not in {
            "true",
            "false",
        }:
            raise AccountingImportError("invalid trade type")
        trade_id = row["trade_id"].strip()
        if not trade_id or trade_id in seen:
            raise AccountingImportError("duplicate or missing trade ID")
        seen.add(trade_id)
        event_at = datetime.fromisoformat(row["order_execution_time"].strip()).replace(tzinfo=UTC)
        if event_at.date() != _date(row["trade_date"], "trade date"):
            raise AccountingImportError("trade date and execution time disagree")
        quantity = _decimal(row["quantity"], "quantity", positive=True)
        price = _decimal(row["price"], "price", nonnegative=True)
        rows.append(
            {
                "symbol": symbol,
                "identity": identity,
                "market": exchange,
                "instrument_type": segment,
                "event_at": event_at,
                "side": row["trade_type"].upper(),
                "quantity": quantity,
                "unit_price": price,
                "gross_amount": quantity * price,
                "source_event_hash": _hash([trade_id, row["order_id"].strip()]),
            }
        )
    return _parsed(raw, "ZERODHA_TRADEBOOK", rows, "event_at")


def parse_zerodha_ledger(path: Path):
    raw, source = _csv(Path(path), ZERODHA_LEDGER_HEADERS)
    if (
        source[0]["particulars"] != "Opening Balance"
        or source[-1]["particulars"] != "Closing Balance"
    ):
        raise AccountingImportError("ledger boundary rows missing")
    opening = _decimal(source[0]["net_balance"], "opening balance")
    closing = _decimal(source[-1]["net_balance"], "closing balance")
    previous = opening
    rows = []
    for index, row in enumerate(source[1:-1], 1):
        day = _date(row["posting_date"], "posting date")
        debit = _decimal(row["debit"] or "0", "debit", nonnegative=True)
        credit = _decimal(row["credit"] or "0", "credit", nonnegative=True)
        balance = _decimal(row["net_balance"], "net balance")
        if debit and credit:
            raise AccountingImportError("ledger row has both debit and credit")
        if previous + credit - debit != balance:
            raise AccountingImportError("ledger balance continuity failed")
        previous = balance
        voucher = row["voucher_type"].strip()
        if voucher not in {"Bank Receipts", "Bank Payments", "Book Voucher", "Journal Entry"}:
            raise AccountingImportError("unknown ledger voucher type")
        rows.append(
            {
                "event_at": datetime.combine(day, datetime.min.time(), UTC),
                "voucher_type": voucher,
                "debit": debit,
                "credit": credit,
                "source_event_hash": _hash([index, *row.values()]),
            }
        )
    if previous != closing:
        raise AccountingImportError("closing balance does not reconcile")
    parsed = _parsed(raw, "ZERODHA_LEDGER", rows, "event_at")
    parsed["opening_balance"] = opening
    parsed["closing_balance"] = closing
    return parsed


def parse_zerodha_dividends(path: Path):
    raw, source = _csv(Path(path), ZERODHA_DIVIDEND_HEADERS)
    rows = []
    seen = set()
    for row in source:
        symbol = row["Symbol"].strip()
        day = _date(row["Ex-date"], "ex-date")
        quantity = _decimal(row["Qty"], "dividend quantity", positive=True)
        rate = _decimal(row["Dividend per share"], "dividend rate", nonnegative=True)
        amount = _decimal(row["Total dividend"], "dividend amount", nonnegative=True)
        if abs(quantity * rate - amount) > Decimal("0.01"):
            raise AccountingImportError("dividend amount does not reconcile")
        natural = [symbol, day, quantity, rate, amount]
        fingerprint = _hash(natural)
        if fingerprint in seen:
            raise AccountingImportError("duplicate dividend row")
        seen.add(fingerprint)
        rows.append(
            {
                "symbol": symbol,
                "event_at": datetime.combine(day, datetime.min.time(), UTC),
                "gross_amount": amount,
                "date_evidence": "SUBSTITUTED_EX_DATE",
                "source_event_hash": fingerprint,
            }
        )
    return _parsed(raw, "ZERODHA_DIVIDENDS", rows, "event_at")


def parse_vested_transactions(path: Path):
    path = Path(path)
    raw = path.read_bytes()
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        raise AccountingImportError(f"invalid XLSX: {type(exc).__name__}") from exc
    if not set(VESTED_SHEETS) <= set(workbook.sheetnames):
        raise AccountingImportError("Vested transaction sheets missing")
    parsed_sheets = {}
    for name, expected in VESTED_SHEETS.items():
        values = list(workbook[name].iter_rows(values_only=True))
        if not values or tuple(str(x).strip() for x in values[0]) != expected:
            raise AccountingImportError(f"{name} headers changed")
        parsed_sheets[name] = [row for row in values[1:] if any(x is not None for x in row)]
    trades = []
    for row in parsed_sheets["Trades"]:
        data = dict(zip(VESTED_SHEETS["Trades"], row, strict=True))
        activity = str(data["Activity"]).strip()
        ticker = str(data["Ticker"]).strip()
        if activity not in {"Buy", "Sell"} or not re.fullmatch(r"[A-Z][A-Z0-9.-]{0,9}", ticker):
            raise AccountingImportError("invalid Vested trade")
        quantity = _decimal(data["Quantity"], "quantity", positive=True)
        price = _decimal(data["Price Per Share (in USD)"], "price", nonnegative=True)
        cash = abs(_decimal(data["Cash Amount (in USD)"], "cash amount"))
        commission = _decimal(data["Commission Charges (in USD)"], "commission", nonnegative=True)
        expected_cash = quantity * price + commission
        if activity == "Sell":
            expected_cash = quantity * price - commission
        if abs(expected_cash - cash) > Decimal("0.02"):
            raise AccountingImportError("Vested trade cash does not reconcile")
        trades.append(
            {
                "symbol": ticker,
                "identity": ticker,
                "market": "US",
                "instrument_type": "EQUITY",
                "event_at": _event_time(data["Date"], data["Time (in UTC)"], "trade"),
                "side": activity.upper(),
                "quantity": quantity,
                "unit_price": price,
                "gross_amount": cash,
                "source_event_hash": _hash(list(row)),
            }
        )
    flows = []
    for row in parsed_sheets["Transfers"]:
        data = dict(zip(VESTED_SHEETS["Transfers"], row, strict=True))
        activity = str(data["Activity"]).strip()
        if activity not in {"Deposit", "Withdrawal"}:
            raise AccountingImportError("invalid Vested transfer")
        flows.append(
            {
                "event_at": _event_time(data["Date"], data["Time (in UTC)"], "transfer"),
                "direction": activity.upper(),
                "amount": abs(_decimal(data["Cash Amount (in USD)"], "transfer", positive=True)),
                "date_evidence": "SOURCE_PAYMENT_DATE",
                "source_event_hash": _hash(list(row)),
            }
        )
    income = []
    for row in parsed_sheets["Income"]:
        data = dict(zip(VESTED_SHEETS["Income"], row, strict=True))
        activity = str(data["Activity"]).strip()
        if activity not in {"Dividend", "Interest", "Tax"}:
            raise AccountingImportError("invalid Vested income type")
        income.append(
            {
                "symbol": str(data["Ticker"] or "").strip() or None,
                "event_at": _event_time(data["Date"], data["Time (in UTC)"], "income"),
                "kind": activity.upper(),
                "amount": abs(_decimal(data["Gross Cash Amount (in USD)"], "income")),
                "date_evidence": "SOURCE_PAYMENT_DATE",
                "source_event_hash": _hash(list(row)),
            }
        )
    fees = []
    all_types = []
    account_balances = []
    for row in parsed_sheets["All Transactions"]:
        data = dict(zip(VESTED_SHEETS["All Transactions"], row, strict=True))
        kind = str(data["Type"]).strip()
        all_types.append(kind)
        balance_at = _event_time(data["Date"], data["Time (in UTC)"], "account balance")
        balance = _decimal(data["Account Balance"], "account balance")
        account_balances.append((balance_at, balance, _hash(list(row))))
        if kind == "FEE":
            fees.append(
                {
                    "event_at": _event_time(data["Date"], data["Time (in UTC)"], "fee"),
                    "amount": abs(_decimal(data["Amount"], "fee", positive=True)),
                    "source_event_hash": _hash(list(row)),
                }
            )
    allowed = {"DIVTAX", "DIV", "SSAL", "SPUR", "INT", "FEE", "CSR", "CSD"}
    if not set(all_types) <= allowed:
        raise AccountingImportError("unknown Vested account transaction type")
    all_counts = Counter(all_types)
    trade_counts = Counter("SPUR" if row["side"] == "BUY" else "SSAL" for row in trades)
    flow_counts = Counter("CSR" if row["direction"] == "DEPOSIT" else "CSD" for row in flows)
    income_counts = Counter(
        {"DIVIDEND": "DIV", "INTEREST": "INT", "TAX": "DIVTAX"}[row["kind"]] for row in income
    )
    expected_counts = trade_counts + flow_counts + income_counts + Counter({"FEE": len(fees)})
    if all_counts != expected_counts:
        raise AccountingImportError("Vested sheet event counts do not reconcile")
    latest_at, latest_balance, latest_hash = max(account_balances, key=lambda item: item[0])
    cash_valuation = {
        "event_at": latest_at,
        "value": latest_balance,
        "source_event_hash": _hash(["CASH_BALANCE", latest_hash]),
    }
    rows = trades + flows + income + fees + [cash_valuation]
    result = _parsed(raw, "VESTED_TRANSACTIONS", rows, "event_at")
    result.update(
        {
            "trades": trades,
            "flows": flows,
            "income": income,
            "fees": fees,
            "cash_valuation": cash_valuation,
        }
    )
    return result


def parse_zerodha_tax_pnl(path: Path):
    path = Path(path)
    raw = path.read_bytes()
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        raise AccountingImportError(f"invalid XLSX: {type(exc).__name__}") from exc
    trade_sheets = [
        name for name in workbook.sheetnames if name.startswith("Tradewise Exits from ")
    ]
    if len(trade_sheets) != 1 or "Equity and Non Equity" not in workbook.sheetnames:
        raise AccountingImportError("tax P&L sheets changed")
    match = re.fullmatch(r"Tradewise Exits from (\d{4}-\d{2}-\d{2})", trade_sheets[0])
    if not match:
        raise AccountingImportError("tax P&L period changed")
    period_start = _date(match.group(1), "tax period start")
    end_dates = []
    for name in workbook.sheetnames:
        found = re.fullmatch(r"Open Positions as of (\d{4}-\d{2}-\d{2})", name)
        if found:
            end_dates.append(_date(found.group(1), "tax period end"))
    if not end_dates:
        raise AccountingImportError("tax P&L end date missing")
    period_end = max(end_dates)
    expected = (
        "Symbol",
        "ISIN",
        "Entry Date",
        "Exit Date",
        "Quantity",
        "Buy Value",
        "Sell Value",
        "Profit",
        "Period of Holding",
        "Fair Market Value",
        "Taxable Profit",
        "Turnover",
        "Brokerage",
        "Exchange Transaction Charges",
        "IPFT",
        "SEBI Charges",
        "CGST",
        "SGST",
        "IGST",
        "Stamp Duty",
        "STT",
    )
    lots = []
    seen_lots = set()
    worksheet = workbook[trade_sheets[0]]
    active = False
    for values in worksheet.iter_rows(min_col=2, max_col=22, values_only=True):
        if tuple(values) == expected:
            active = True
            continue
        if not active or values[0] in (None, ""):
            continue
        if values[0] == "Symbol":
            active = False
            continue
        if values[2] is None or values[3] is None:
            continue
        symbol = str(values[0]).strip()
        identity = str(values[1] or f"SYMBOL:NSE:{symbol}").strip()
        quantity = _decimal(values[4], "tax lot quantity", positive=True)
        cost = _decimal(values[5], "tax lot cost", nonnegative=True)
        proceeds = _decimal(values[6], "tax lot proceeds", nonnegative=True)
        profit = _decimal(values[7], "tax lot profit")
        if abs((proceeds - cost) - profit) > Decimal("0.02"):
            raise AccountingImportError("tax lot P&L does not reconcile")
        source_event_hash = _hash(list(values))
        if source_event_hash in seen_lots:
            continue
        seen_lots.add(source_event_hash)
        lots.append(
            {
                "symbol": symbol,
                "identity": identity,
                "market": "IN",
                "instrument_type": "EQUITY",
                "acquired_date": _date(values[2], "entry date"),
                "disposed_date": _date(values[3], "exit date"),
                "quantity": quantity,
                "cost_basis": cost,
                "proceeds": proceeds,
                "realized_pnl": profit,
                "event_at": datetime.combine(
                    _date(values[3], "exit date"), datetime.min.time(), UTC
                ),
                "source_event_hash": source_event_hash,
            }
        )
    if not lots:
        raise AccountingImportError("tax P&L has no realized lots")
    summaries = []
    summary_sheet = workbook["Equity and Non Equity"]
    in_accounts = False
    for row in summary_sheet.iter_rows(min_col=2, max_col=3, values_only=True):
        if row == ("Account Head", "Amount"):
            in_accounts = True
            continue
        if in_accounts and row[0] in (None, ""):
            if summaries:
                break
            continue
        if in_accounts:
            amount = _decimal(row[1], "summary amount")
            summaries.append(
                {
                    "segment": "EQUITY_AND_NON_EQUITY",
                    "period_start": period_start,
                    "period_end": period_end,
                    "summary_type": str(row[0]).strip(),
                    "amount": amount,
                    "event_at": datetime.combine(period_end, datetime.min.time(), UTC),
                    "source_event_hash": _hash([period_start, period_end, *row]),
                }
            )
    result = _parsed(raw, "ZERODHA_TAX_PNL", lots + summaries, "event_at")
    result["coverage_start"] = period_start
    result["coverage_end"] = period_end
    result.update({"lots": lots, "summaries": summaries})
    return result


def parse_vested_holdings(path: Path):
    path = Path(path)
    raw = path.read_bytes()
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        raise AccountingImportError(f"invalid XLSX: {type(exc).__name__}") from exc
    if not {"User Details", "Summary", "Holdings"} <= set(workbook.sheetnames):
        raise AccountingImportError("Vested holdings sheets missing")
    user = list(workbook["User Details"].iter_rows(values_only=True))
    allowed_user_headers = {
        ("Period", "User", "Govt ID", "DW Account Number", "Email"),
        ("Period", "User", "Govt Id", "DriveWealth Account Number", "Email"),
    }
    if not user or tuple(user[0]) not in allowed_user_headers:
        raise AccountingImportError("Vested User Details headers changed")
    match = re.fullmatch(r"As of (\d{1,2} [A-Za-z]{3} \d{4})", str(user[1][0]))
    if not match:
        raise AccountingImportError("Vested holdings period changed")
    snapshot = datetime.strptime(match.group(1), "%d %b %Y").date()
    worksheet = workbook["Holdings"]
    values = list(worksheet.iter_rows(values_only=True))
    if not values or tuple(values[0]) != vested.HEADERS:
        raise AccountingImportError("Vested Holdings headers changed")
    valuations = []
    seen = set()
    for row in values[1:]:
        if not any(value is not None for value in row):
            continue
        ticker = str(row[1]).strip()
        if not re.fullmatch(r"[A-Z][A-Z0-9.-]{0,9}", ticker) or ticker in seen:
            raise AccountingImportError("invalid or duplicate Vested holding")
        seen.add(ticker)
        value = _decimal(row[4], "holding value", nonnegative=True)
        cost_basis = _decimal(row[6], "holding cost basis", nonnegative=True)
        valuations.append(
            {
                "symbol": ticker,
                "identity": ticker,
                "market": "US",
                "instrument_type": "EQUITY",
                "event_at": datetime.combine(snapshot, datetime.min.time(), UTC),
                "value": value,
                "cost_basis": cost_basis,
                "source_event_hash": _hash([snapshot, ticker, value, cost_basis]),
            }
        )
    summary = list(workbook["Summary"].iter_rows(values_only=True))
    if not summary or tuple(summary[0]) != (
        "Current Equity Value (USD)",
        "Total Amount Invested (USD)",
        "Investment Returns (USD)",
        "Investment Returns (%)",
    ):
        raise AccountingImportError("Vested Summary headers changed")
    total = _decimal(summary[1][0], "summary value", nonnegative=True)
    if abs(sum(row["value"] for row in valuations) - total) > Decimal("0.02"):
        raise AccountingImportError("Vested holdings do not reconcile")
    result = _parsed(raw, "VESTED_HOLDINGS", valuations, "event_at")
    result["valuations"] = valuations
    return result


def _parsed(raw: bytes, source_type: str, rows: list[dict], date_key: str):
    if not rows:
        raise AccountingImportError("source has no importable rows")
    canonical = _hash(rows)
    dates = [row[date_key].date() for row in rows]
    return {
        "source_type": source_type,
        "content_sha256": hashlib.sha256(raw).hexdigest(),
        "source_fingerprint": canonical,
        "coverage_start": min(dates),
        "coverage_end": max(dates),
        "row_count": len(rows),
        "rows": rows,
    }


def _instrument(conn, account_id, row, currency):
    source_hash = _hash([row["market"], row["identity"]])
    instrument_id = _hash([account_id, source_hash])[:24]
    conn.execute(
        "INSERT INTO portfolio_instrument VALUES (?,?,?,?,?,?) ON CONFLICT DO NOTHING",
        [
            instrument_id,
            row["market"],
            row["symbol"],
            row["instrument_type"],
            currency,
            source_hash,
        ],
    )
    return instrument_id


def store(conn: duckdb.DuckDBPyConnection, account_id: str, parsed: dict, now=None):
    now = now or datetime.now(UTC)
    import_id = _hash([account_id, parsed["source_type"], parsed["content_sha256"]])[:24]
    existing = conn.execute(
        "SELECT import_id FROM accounting_import_run "
        "WHERE account_id=? AND source_type=? AND content_sha256=?",
        [account_id, parsed["source_type"], parsed["content_sha256"]],
    ).fetchone()
    if existing:
        return {"status": "duplicate", "import_id": existing[0], "rows": parsed["row_count"]}
    account = conn.execute(
        "SELECT provider,native_currency FROM portfolio_account WHERE account_id=?", [account_id]
    ).fetchone()
    if not account:
        raise AccountingImportError("unknown portfolio account")
    provider, currency = account
    if not parsed["source_type"].startswith(provider):
        raise AccountingImportError("source provider does not match account")
    conn.execute("BEGIN")
    try:
        conn.execute(
            "INSERT INTO accounting_import_run VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            [
                import_id,
                account_id,
                parsed["source_type"],
                parsed["content_sha256"],
                parsed["coverage_start"],
                parsed["coverage_end"],
                parsed["row_count"],
                parsed["source_fingerprint"],
                now,
                None,
                None,
            ],
        )
        if parsed["source_type"] == "ZERODHA_TRADEBOOK":
            _store_trades(conn, account_id, import_id, parsed["rows"], currency)
        elif parsed["source_type"] == "ZERODHA_LEDGER":
            _store_ledger(conn, account_id, import_id, parsed["rows"], currency)
        elif parsed["source_type"] == "ZERODHA_DIVIDENDS":
            _store_dividends(conn, account_id, import_id, parsed["rows"], currency)
        elif parsed["source_type"] == "VESTED_TRANSACTIONS":
            _store_trades(conn, account_id, import_id, parsed["trades"], currency)
            _store_flows(conn, account_id, import_id, parsed["flows"], currency)
            _store_vested_income(conn, account_id, import_id, parsed["income"], currency)
            _store_fees(conn, account_id, import_id, parsed["fees"], currency)
            _store_account_valuation(
                conn, account_id, import_id, parsed["cash_valuation"], currency
            )
        elif parsed["source_type"] == "ZERODHA_TAX_PNL":
            _store_tax_lots(conn, account_id, import_id, parsed["lots"], currency)
            _store_summaries(conn, account_id, import_id, parsed["summaries"], currency)
        elif parsed["source_type"] == "VESTED_HOLDINGS":
            _store_valuations(conn, account_id, import_id, parsed["valuations"], currency)
        else:
            raise AccountingImportError("unsupported source type")
        conn.execute("COMMIT")
    except Exception:
        conn.execute("ROLLBACK")
        raise
    return {"status": "stored", "import_id": import_id, "rows": parsed["row_count"]}


def _store_trades(conn, account_id, import_id, rows, currency):
    for row in rows:
        instrument_id = _instrument(conn, account_id, row, currency)
        event_id = _hash([account_id, "TRADE", row["source_event_hash"]])[:24]
        conn.execute(
            "INSERT INTO portfolio_transaction VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            [
                event_id,
                import_id,
                account_id,
                instrument_id,
                row["event_at"],
                row["side"],
                row["quantity"],
                row["unit_price"],
                row["gross_amount"],
                currency,
                row["source_event_hash"],
            ],
        )


def _store_ledger(conn, account_id, import_id, rows, currency):
    flows = []
    fees = []
    for row in rows:
        if row["voucher_type"] in {"Bank Receipts", "Bank Payments"}:
            flows.append(
                {
                    "event_at": row["event_at"],
                    "direction": "DEPOSIT" if row["credit"] else "WITHDRAWAL",
                    "amount": row["credit"] or row["debit"],
                    "date_evidence": "SOURCE_POSTING_DATE",
                    "source_event_hash": row["source_event_hash"],
                }
            )
        elif row["voucher_type"] == "Journal Entry":
            fees.append(
                {
                    "event_at": row["event_at"],
                    "amount": row["debit"],
                    "source_event_hash": row["source_event_hash"],
                }
            )
    _store_flows(conn, account_id, import_id, flows, currency)
    _store_fees(conn, account_id, import_id, fees, currency)


def _store_flows(conn, account_id, import_id, rows, currency):
    for row in rows:
        event_id = _hash([account_id, "FLOW", row["source_event_hash"]])[:24]
        conn.execute(
            "INSERT INTO portfolio_cash_flow VALUES (?,?,?,?,?,?,?,?,?)",
            [
                event_id,
                import_id,
                account_id,
                row["event_at"],
                row["direction"],
                row["amount"],
                currency,
                row["date_evidence"],
                row["source_event_hash"],
            ],
        )


def _store_fees(conn, account_id, import_id, rows, currency):
    for row in rows:
        event_id = _hash([account_id, "FEE", row["source_event_hash"]])[:24]
        conn.execute(
            "INSERT INTO portfolio_fee VALUES (?,?,?,?,?,?,?,?)",
            [
                event_id,
                import_id,
                account_id,
                row["event_at"],
                "BROKER_REPORTED",
                row["amount"],
                currency,
                row["source_event_hash"],
            ],
        )


def _store_dividends(conn, account_id, import_id, rows, currency):
    for row in rows:
        event_id = _hash([account_id, "DIVIDEND", row["source_event_hash"]])[:24]
        existing = conn.execute(
            "SELECT source_event_hash FROM portfolio_income WHERE event_id=?", [event_id]
        ).fetchone()
        if existing:
            if existing[0] != row["source_event_hash"]:
                raise AccountingImportError("dividend identity conflict")
            continue
        conn.execute(
            "INSERT INTO portfolio_income VALUES (?,?,?,?,?,?,?,?,?,?)",
            [
                event_id,
                import_id,
                account_id,
                None,
                row["event_at"],
                "DIVIDEND",
                row["gross_amount"],
                currency,
                row["date_evidence"],
                row["source_event_hash"],
            ],
        )


def _store_vested_income(conn, account_id, import_id, rows, currency):
    for row in rows:
        event_id = _hash([account_id, row["kind"], row["source_event_hash"]])[:24]
        if row["kind"] == "TAX":
            conn.execute(
                "INSERT INTO portfolio_tax VALUES (?,?,?,?,?,?,?,?)",
                [
                    event_id,
                    import_id,
                    account_id,
                    row["event_at"],
                    "WITHHOLDING",
                    row["amount"],
                    currency,
                    row["source_event_hash"],
                ],
            )
        else:
            conn.execute(
                "INSERT INTO portfolio_income VALUES (?,?,?,?,?,?,?,?,?,?)",
                [
                    event_id,
                    import_id,
                    account_id,
                    None,
                    row["event_at"],
                    row["kind"],
                    row["amount"],
                    currency,
                    row["date_evidence"],
                    row["source_event_hash"],
                ],
            )


def _store_tax_lots(conn, account_id, import_id, rows, currency):
    for row in rows:
        instrument_id = _instrument(conn, account_id, row, currency)
        lot_id = _hash([account_id, "LOT", row["source_event_hash"]])[:24]
        conn.execute(
            "INSERT INTO portfolio_tax_lot VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                lot_id,
                import_id,
                account_id,
                instrument_id,
                row["acquired_date"],
                row["disposed_date"],
                row["quantity"],
                row["cost_basis"],
                row["proceeds"],
                row["realized_pnl"],
                "REALIZED",
                row["source_event_hash"],
            ],
        )


def _store_summaries(conn, account_id, import_id, rows, currency):
    for row in rows:
        summary_id = _hash([account_id, "SUMMARY", row["source_event_hash"]])[:24]
        conn.execute(
            "INSERT INTO broker_reported_summary VALUES (?,?,?,?,?,?,?,?,?,?)",
            [
                summary_id,
                import_id,
                account_id,
                row["segment"],
                row["period_start"],
                row["period_end"],
                row["summary_type"],
                row["amount"],
                currency,
                row["source_event_hash"],
            ],
        )


def _store_valuations(conn, account_id, import_id, rows, currency):
    for row in rows:
        instrument_id = _instrument(conn, account_id, row, currency)
        valuation_id = _hash([account_id, "VALUATION", row["source_event_hash"]])[:24]
        conn.execute(
            "INSERT INTO portfolio_valuation VALUES (?,?,?,?,?,?,?,?,?,?)",
            [
                valuation_id,
                import_id,
                account_id,
                instrument_id,
                row["event_at"],
                row["value"],
                row["cost_basis"],
                currency,
                "SOURCE_SNAPSHOT",
                row["source_event_hash"],
            ],
        )


def _store_account_valuation(conn, account_id, import_id, row, currency):
    valuation_id = _hash([account_id, "CASH_VALUATION", row["source_event_hash"]])[:24]
    conn.execute(
        "INSERT INTO portfolio_valuation VALUES (?,?,?,?,?,?,?,?,?,?)",
        [
            valuation_id,
            import_id,
            account_id,
            None,
            row["event_at"],
            row["value"],
            row["value"],
            currency,
            "SOURCE_SNAPSHOT",
            row["source_event_hash"],
        ],
    )
