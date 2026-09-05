"""Workbook export v2 (T1.6b): adapted to the original research MFanalysis2.xlsx format.

Single sheet "Sheet 1", headers in row 1 (A-Q identical labels to
myvault/MFanalysis2.xlsx), values in percent points (returns/SD as 11.57,
capture ratios as 90), E column as '+x.x (O)' style without % sign,
'Average' instead of 'At Category'. Columns R/S added on top of his layout:
Window = lookback actually used per row ('3Y', fallback '1Y', else blank)
and Sharpe. Conditional formatting replicas of MFanalysis2: color scales on
D/F/L/O and a full-sheet red fill for cells containing 'Underperformer'.

Per fund the longest available lookback wins (3Y, else 1Y); young funds stay
blank rather than mislabeling windows. Category stats use min_category_peers
from config/metrics.json (now 2).
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from invest import db

log = logging.getLogger("invest.export")

LOOKBACKS = ["3Y", "1Y"]  # preference order; longest available wins

HEADERS = [
    "Fund Name",
    "Category",
    "Category Avg (3Y)",
    "Fund Returns (3Y)",
    "Outperformer/ Under",
    "SD (Volatility)",
    "Category SD",
    "Volatility Classification",
    "Fund's Beta (vs Market)",
    "Category Beta",
    "Risk Profile",
    "Upside CR (Bull Markets)",
    "Category Upside CR",
    "Upside Result",
    "Downside CR (Bear Markets)",
    "Category Downside CR",
    "Downside Result",
    # additions beyond MFanalysis2, kept last so A-Q stays byte-compatible:
    "Window",
    "Sharpe",
]

RESULT_MAP = {
    "Outperformer": "Outperformer",
    "Underperformer": "Underperformer",
    "At Category": "Average",
}


def category_label(raw: str | None) -> str | None:
    """'Equity Scheme - Mid Cap Fund' -> 'Mid Cap'; None passthrough."""
    if not raw:
        return None
    tail = raw.split("-")[-1].strip()
    for suffix in (" Funds", " Fund"):
        if tail.endswith(suffix):
            tail = tail[: -len(suffix)]
    return tail or None


def _diff_string(result: str | None, fund: float | None, cat: float | None) -> str | None:
    """MFanalysis2 convention: '0.4 (O)' / '-17.7 (U)' in percentage points."""
    if result is None or fund is None or cat is None:
        return None
    marker = RESULT_MAP[result]
    if marker == "Average":
        return "0.0"
    return f"{(fund - cat) * 100:.1f} ({'O' if marker == 'Outperformer' else 'U'})"


def _fetch(conn) -> list[dict]:
    """Tracked funds joined to best-available lookback metrics (3Y > 1Y)."""
    out = []
    schemes = conn.execute(
        "SELECT scheme_code, display_name, category FROM mf_scheme "
        "WHERE display_name IS NOT NULL ORDER BY display_name"
    ).fetchall()
    for code, name, cat in schemes:
        row = {"name": name, "category": cat, "window": None}
        for lb in LOOKBACKS:
            ret = conn.execute(
                "SELECT fund_return, category_avg_return, result FROM mf_return_metrics "
                "WHERE scheme_code=? AND lookback=?",
                [code, lb],
            ).fetchone()
            risk = conn.execute(
                "SELECT sd, category_sd, volatility_class, beta, category_beta, "
                "risk_profile, upside_cr, category_upside_cr, upside_result, "
                "downside_cr, category_downside_cr, downside_result, sharpe "
                "FROM mf_risk_metrics WHERE scheme_code=? AND lookback=? "
                "ORDER BY calculated_at DESC LIMIT 1",
                [code, lb],
            ).fetchone()
            if ret and risk:
                keys = ["fund_return", "cat_return", "result"]
                row.update(dict(zip(keys, ret, strict=True)))
                rkeys = [
                    "sd",
                    "cat_sd",
                    "vol_class",
                    "beta",
                    "cat_beta",
                    "risk_profile",
                    "up_cr",
                    "cat_up_cr",
                    "up_result",
                    "down_cr",
                    "cat_down_cr",
                    "down_result",
                    "sharpe",
                ]
                row.update(dict(zip(rkeys, risk, strict=True)))
                row["window"] = lb
                break
        out.append(row)
    return out


def _add_conditional_formatting(ws) -> None:
    """Replicas of MFanalysis2.xlsx rules (extracted 2026-08-25)."""
    rYGW = ("FFF8696B", "FFFFEB84", "FF63BE7B")  # red-yellow-green
    ws.conditional_formatting.add(
        "D2:D1048576",
        ColorScaleRule(
            start_type="percentile",
            start_value=0,
            start_color=rYGW[0],
            mid_type="num",
            mid_value=13,
            mid_color=rYGW[1],
            end_type="percentile",
            end_value=100,
            end_color=rYGW[2],
        ),
    )
    ws.conditional_formatting.add(
        "D2:D1048576",
        ColorScaleRule(
            start_type="percentile",
            start_value=0,
            start_color=rYGW[0],
            mid_type="num",
            mid_value=50,
            mid_color=rYGW[1],
            end_type="percentile",
            end_value=100,
            end_color=rYGW[2],
        ),
    )
    # low SD good -> green at the low end (reversed scale)
    ws.conditional_formatting.add(
        "F2:F1048576",
        ColorScaleRule(
            start_type="percentile",
            start_value=0,
            start_color=rYGW[2],
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFCFCFF",
            end_type="percentile",
            end_value=100,
            end_color=rYGW[0],
        ),
    )
    ws.conditional_formatting.add(
        "L2:L1048576",
        ColorScaleRule(
            start_type="num",
            start_value=10,
            start_color="FFFFC000",
            end_type="num",
            end_value=90,
            end_color="FF00B050",
        ),
    )
    ws.conditional_formatting.add(
        "L2:L1048576",
        ColorScaleRule(
            start_type="percentile",
            start_value=0,
            start_color=rYGW[0],
            mid_type="num",
            mid_value=50,
            mid_color="FFFCFCFF",
            end_type="percentile",
            end_value=100,
            end_color=rYGW[2],
        ),
    )
    ws.conditional_formatting.add(
        "O2:O1048576",
        ColorScaleRule(
            start_type="min", start_color="FF92D050", end_type="max", end_color="FF548235"
        ),
    )
    dxf = DifferentialStyle(fill=PatternFill(bgColor="FFFF0000"))
    rule = Rule(
        type="containsText",
        operator="containsText",
        text="Underperformer",
        formula=['NOT(ISERROR(SEARCH("Underperformer",A1)))'],
        dxf=dxf,
    )
    ws.conditional_formatting.add("A1:XFD1048576", rule)
    # negative Sharpe -> red font (risk-adjusted under water)
    ws.conditional_formatting.add(
        "S2:S1048576", CellIsRule(operator="lessThan", formula=["0"], font=Font(color="FFCC0000"))
    )


def _sorted_rows(rows: list[dict]) -> list[dict]:
    """Category blocks A→Z, best 3Y return first inside each block, no-data last."""
    return sorted(
        rows,
        key=lambda r: (
            category_label(r.get("category")) or "~",
            -(r.get("fund_return") if r.get("fund_return") is not None else -1.0),
            r.get("name") or "",
        ),
    )


_THIN = Side(style="thin", color="FFBFBFBF")
_MEDIUM = Side(style="medium", color="FF808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WINDOW_FILL = {
    "3Y": PatternFill("solid", fgColor="FFC6EFCE"),
    "1Y": PatternFill("solid", fgColor="FFFFEB9C"),
}


def _pp(v):
    """Fraction -> percent points, 2dp (sheet convention)."""
    return None if v is None else round(v * 100, 2)


def _row_values(row: dict) -> list:
    """Transform one fetched fund dict into the sheet's 19 cell values.

    Shared by _write_sheet and verify.py so the sheet can never drift from
    what the verifier expects.
    """
    return [
        row["name"],
        category_label(row.get("category")),
        _pp(row.get("cat_return")),
        _pp(row.get("fund_return")),
        _diff_string(row.get("result"), row.get("fund_return"), row.get("cat_return")),
        _pp(row.get("sd")),
        _pp(row.get("cat_sd")),
        row.get("vol_class"),
        row.get("beta"),
        row.get("cat_beta"),
        row.get("risk_profile"),
        _pp(row.get("up_cr")),
        _pp(row.get("cat_up_cr")),
        RESULT_MAP.get(row.get("up_result") or "", None),
        _pp(row.get("down_cr")),
        _pp(row.get("cat_down_cr")),
        RESULT_MAP.get(row.get("down_result") or "", None),
        row.get("window"),
        None if row.get("sharpe") is None else round(row["sharpe"], 2),
    ]


def _write_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Sheet 1"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF305496")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_MEDIUM)
    ws.row_dimensions[1].height = 30

    pct_cols, num_cols, cr_cols = {3, 4}, {6, 7}, {12, 13, 15, 16}
    rows = _sorted_rows(rows)
    prev_cat = None
    for r_i, row in enumerate(rows, start=2):
        has_data = row.get("window") is not None
        cat = category_label(row.get("category"))
        new_block = cat != prev_cat
        prev_cat = cat
        values = _row_values(row)
        muted = Font(color="FF9C9C9C", italic=True) if not has_data else None
        top_side = _MEDIUM if new_block else _THIN
        for c_i, v in enumerate(values, start=1):
            cell = ws.cell(row=r_i, column=c_i)
            cell.border = Border(left=_THIN, right=_THIN, top=top_side, bottom=_THIN)
            if muted:
                cell.font = muted
            if v is None:
                continue
            cell.value = v
            if c_i == 18 and row.get("window") in _WINDOW_FILL:  # window chip
                cell.fill = _WINDOW_FILL[row["window"]]
                cell.alignment = Alignment(horizontal="center")
            elif c_i in pct_cols:
                cell.number_format = "0.00"
            elif c_i in num_cols:
                cell.number_format = "0.00"
            elif c_i in cr_cols:
                cell.number_format = "0"
            elif c_i in (9, 10, 19):
                cell.number_format = "0.00"

    widths = [42, 16, 11, 11, 13, 10, 10, 13, 10, 10, 12, 10, 10, 13, 10, 11, 13, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    _add_conditional_formatting(ws)


def export_workbook(conn, out_path: str | Path, data_end=None) -> dict:
    """Write the adapted single-sheet workbook; returns summary."""
    summary = {"path": str(out_path)}
    (summary["funds"],) = conn.execute(
        "SELECT count(*) FROM mf_scheme WHERE display_name IS NOT NULL"
    ).fetchone()
    if not summary["funds"]:
        msg = "no tracked funds; run ingest first"
        raise RuntimeError(msg)
    if not data_end:
        (data_end,) = conn.execute("SELECT max(nav_date) FROM mf_nav").fetchone()
    summary["data_end"] = data_end

    rows = _fetch(conn)
    summary["computed_3y"] = sum(1 for r in rows if r["window"] == "3Y")
    summary["computed_1y"] = sum(1 for r in rows if r["window"] == "1Y")
    wb = Workbook()
    _write_sheet(wb, rows)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    check = load_workbook(out, read_only=True)  # opens-clean self-check
    assert check.sheetnames == ["Sheet 1"]
    check.close()
    summary["rows"] = len(rows)
    log.info(
        "exported %s: %d funds (%d 3Y, %d 1Y-only), data end %s",
        out,
        len(rows),
        summary["computed_3y"],
        summary["computed_1y"],
        data_end,
    )
    return summary


def main(argv: list[str] | None = None) -> int:
    import argparse

    logging.basicConfig(
        stream=sys.stderr,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    parser = argparse.ArgumentParser(prog="invest-export")
    parser.add_argument("--db", default="data/invest.duckdb")
    parser.add_argument(
        "--out", default=None, help="default data/exports/MFs_export_<data-end>.xlsx"
    )
    args = parser.parse_args(argv)

    conn = db.connect(args.db)
    db.init_schema(conn)
    (data_end,) = conn.execute("SELECT max(nav_date) FROM mf_nav").fetchone()
    out = args.out or f"data/exports/MFs_export_{data_end}.xlsx"
    s = export_workbook(conn, out, data_end)
    print(
        f"wrote {s['path']} — {s['rows']} funds "
        f"({s['computed_3y']} 3Y + {s['computed_1y']} 1Y-only), data through {data_end}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
