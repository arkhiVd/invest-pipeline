"""Strict, privacy-safe import of operator-exported Vested holdings XLSX."""

from __future__ import annotations

import hashlib
import json
import math
import re
import sys
from datetime import UTC
from datetime import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

from invest import db, watchlist

HEADERS = (
    "Name",
    "Ticker",
    "Total Shares Held",
    "Current Price (USD)",
    "Current Value (USD)",
    "Average Cost (USD)",
    "Total Amount Invested (USD)",
    "Investment Returns (USD)",
    "Investment Returns (%)",
    "Daily Change (USD)",
    "Daily Change (%)",
)
FIELDS = (
    "name",
    "ticker",
    "quantity",
    "current_price_usd",
    "current_value_usd",
    "average_cost_usd",
    "invested_usd",
    "return_usd",
    "return_pct",
)


class VestedError(RuntimeError):
    pass


def _pct(v, label):
    if isinstance(v, str) and re.fullmatch(r"-?\d+(?:\.\d+)?%", v.strip()):
        return float(v.strip()[:-1])
    return _num(v, label)


def _num(v, label):
    if isinstance(v, str) and re.fullmatch(r"-?\d+(?:\.\d+)?", v.strip()):
        v = float(v.strip())
    if isinstance(v, bool) or not isinstance(v, (int, float)) or not math.isfinite(float(v)):
        raise VestedError(f"{label} must be finite numeric")
    return float(v)


def parse(path: Path):
    raw = path.read_bytes()
    source = hashlib.sha256(raw).hexdigest()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise VestedError(f"invalid XLSX: {type(exc).__name__}") from exc
    if not {"User Details", "Summary", "Holdings"} <= set(wb.sheetnames):
        raise VestedError("required sheets are missing")
    user = wb["User Details"]
    expected_user = ("Period", "User", "Govt ID", "DW Account Number", "Email")
    if tuple(cell.value for cell in user[1]) != expected_user:
        raise VestedError("User Details headers changed")
    period = user["A2"].value
    m = re.fullmatch(r"As of (\d{1,2} [A-Za-z]{3} \d{4})", period or "")
    if not m:
        raise VestedError("invalid snapshot period")
    snapshot = dt.strptime(m.group(1), "%d %b %Y").date()
    ws = wb["Holdings"]
    if tuple(c.value for c in ws[1]) != HEADERS:
        raise VestedError("Holdings headers changed")
    rows = []
    seen = set()
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in values):
            continue
        name, ticker = values[:2]
        if (
            not isinstance(name, str)
            or not name.strip()
            or not isinstance(ticker, str)
            or not re.fullmatch(r"[A-Z][A-Z0-9.-]{0,9}", ticker)
        ):
            raise VestedError("invalid holding identity")
        if ticker in seen:
            raise VestedError("duplicate ticker")
        seen.add(ticker)
        nums = [_num(values[i], HEADERS[i]) for i in range(2, 11)]
        if nums[0] <= 0 or min(nums[1], nums[2], nums[3], nums[4]) < 0:
            raise VestedError("holding values outside range")
        rows.append(dict(zip(FIELDS, [name.strip(), ticker, *nums[:7]], strict=True)))
    if not rows:
        raise VestedError("no holdings")
    summary = wb["Summary"]
    expected = (
        "Current Equity Value (USD)",
        "Total Amount Invested (USD)",
        "Investment Returns (USD)",
        "Investment Returns (%)",
    )
    if tuple(c.value for c in summary[1]) != expected:
        raise VestedError("Summary headers changed")
    total_value = _num(summary["A2"].value, "summary value")
    invested = _num(summary["B2"].value, "summary invested")
    total_return = _num(summary["C2"].value, "summary return")
    return_pct = _pct(summary["D2"].value, "summary return percentage")
    derived_pct = (total_value - invested) / invested * 100 if invested else 0.0
    if (
        abs(sum(r["current_value_usd"] for r in rows) - total_value) > 0.02
        or abs(sum(r["invested_usd"] for r in rows) - invested) > 0.02
        or abs(sum(r["return_usd"] for r in rows) - total_return) > 0.02
        or abs(total_value - invested - total_return) > 0.02
        or abs(derived_pct - return_pct) > 0.02
    ):
        raise VestedError("Summary totals do not reconcile")
    rows.sort(key=lambda row: row["ticker"])
    canonical = json.dumps(rows, sort_keys=True, separators=(",", ":"))
    return (
        snapshot,
        source,
        hashlib.sha256(canonical.encode()).hexdigest(),
        rows,
        total_value,
        invested,
    )


def store(conn, path, now=None):
    now = now or dt.now(UTC)
    snapshot, source, content, rows, value, invested = parse(Path(path))
    run = hashlib.sha256(f"vested:{snapshot}:{content}".encode()).hexdigest()[:24]
    existing = conn.execute(
        "select run_id from vested_snapshot_run where snapshot_date=? and content_sha256=?",
        [snapshot, content],
    ).fetchone()
    if existing:
        return {"run_id": existing[0], "status": "duplicate", "holdings": len(rows)}
    conn.execute("begin")
    try:
        conn.execute(
            "insert into vested_snapshot_run values (?, 'vested', ?, ?, ?, ?, ?, ?, ?)",
            [run, snapshot, source, content, len(rows), value, invested, now],
        )
        for r in rows:
            conn.execute(
                "insert into vested_holding "
                "(run_id, name, ticker, quantity, current_price_usd, current_value_usd, "
                "average_cost_usd, invested_usd, return_usd, return_pct) "
                "values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [run, *[r[k] for k in FIELDS]],
            )
        conn.execute("commit")
    except Exception:
        conn.execute("rollback")
        raise
    return {
        "run_id": run,
        "status": "stored",
        "holdings": len(rows),
        "snapshot_date": str(snapshot),
        "current_value_usd": value,
        "invested_usd": invested,
    }


def integrity(conn, run):
    p = conn.execute(
        "select content_sha256, holding_count, current_value_usd, invested_usd "
        "from vested_snapshot_run where run_id=?",
        [run],
    ).fetchone()
    if not p:
        return False
    rows = [
        dict(zip(FIELDS, r, strict=True))
        for r in conn.execute(
            "select " + ",".join(FIELDS) + " from vested_holding where run_id=? order by ticker",
            [run],
        ).fetchall()
    ]
    return (
        len(rows) == p[1]
        and abs(sum(row["current_value_usd"] for row in rows) - p[2]) <= 0.02
        and abs(sum(row["invested_usd"] for row in rows) - p[3]) <= 0.02
        and hashlib.sha256(
            json.dumps(rows, sort_keys=True, separators=(",", ":")).encode()
        ).hexdigest()
        == p[0]
    )


def main(argv=None):
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--db", default="data/invest.duckdb")
    ap.add_argument("--out", type=Path)
    a = ap.parse_args(argv)
    c = db.connect(a.db)
    try:
        db.init_schema(c)
        result = store(c, a.xlsx)
        text = (
            f"VESTED IMPORT status={result['status']} run={result['run_id']} "
            f"holdings={result['holdings']}\n"
            "Private identifiers excluded. No trade instruction is produced."
        )
        watchlist.atomic_write(str(a.out), text) if a.out else None
        print(text)
        return 0
    except (OSError, VestedError, ValueError) as e:
        print(f"Vested import failed: {e}", file=sys.stderr)
        return 1
    finally:
        c.close()


if __name__ == "__main__":
    raise SystemExit(main())
