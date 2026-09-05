"""T11.3 strict accounting importer fixtures on disposable databases."""

import csv
from datetime import UTC, datetime
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook

from invest import accounting, accounting_import, db


@pytest.fixture()
def conn():
    c = duckdb.connect()
    db.init_schema(c)
    accounting.install_schema(c)
    c.executemany(
        "INSERT INTO portfolio_account VALUES (?,?,?,?,?,?)",
        [
            ("zerodha", "ZERODHA", "EQUITY", "INR", "UNPROVEN", datetime.now(UTC)),
            ("vested", "VESTED", "US_EQUITY", "USD", "UNPROVEN", datetime.now(UTC)),
        ],
    )
    yield c
    c.close()


def csv_file(path: Path, headers, rows):
    with path.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def test_zerodha_tradebook_strict_store_and_replay(conn, tmp_path):
    path = tmp_path / "trade.csv"
    row = [
        "ACME",
        "IDENTITY",
        "2026-01-02",
        "NSE",
        "EQ",
        "EQ",
        "buy",
        "false",
        "2",
        "10.25",
        "trade-1",
        "order-1",
        "2026-01-02 10:11:12",
    ]
    csv_file(path, accounting_import.ZERODHA_TRADE_HEADERS, [row])
    parsed = accounting_import.parse_zerodha_tradebook(path)
    first = accounting_import.store(conn, "zerodha", parsed)
    second = accounting_import.store(conn, "zerodha", parsed)
    assert first["status"] == "stored"
    assert second["status"] == "duplicate"
    assert conn.execute("SELECT quantity,gross_amount FROM portfolio_transaction").fetchone() == (
        pytest.approx(2),
        pytest.approx(20.5),
    )
    csv_file(path, accounting_import.ZERODHA_TRADE_HEADERS, [row, row])
    with pytest.raises(accounting_import.AccountingImportError, match="duplicate"):
        accounting_import.parse_zerodha_tradebook(path)


def test_zerodha_ledger_requires_exact_rollforward(conn, tmp_path):
    path = tmp_path / "ledger.csv"
    rows = [
        ["Opening Balance", "", "", "", "", "", "0"],
        ["deposit", "2026-01-01", "NSE-EQ", "Bank Receipts", "0", "100", "100"],
        ["charge", "2026-01-02", "NSE-EQ", "Journal Entry", "2", "0", "98"],
        ["withdraw", "2026-01-03", "NSE-EQ", "Bank Payments", "50", "0", "48"],
        ["Closing Balance", "", "", "", "", "", "48"],
    ]
    csv_file(path, accounting_import.ZERODHA_LEDGER_HEADERS, rows)
    parsed = accounting_import.parse_zerodha_ledger(path)
    accounting_import.store(conn, "zerodha", parsed)
    assert conn.execute(
        "SELECT direction,amount FROM portfolio_cash_flow ORDER BY event_at"
    ).fetchall() == [
        ("DEPOSIT", pytest.approx(100)),
        ("WITHDRAWAL", pytest.approx(50)),
    ]
    assert conn.execute("SELECT amount FROM portfolio_fee").fetchone()[0] == pytest.approx(2)
    rows[2][-1] = "99"
    csv_file(path, accounting_import.ZERODHA_LEDGER_HEADERS, rows)
    with pytest.raises(accounting_import.AccountingImportError, match="continuity"):
        accounting_import.parse_zerodha_ledger(path)


def test_dividend_ex_date_is_mandatory_estimate_label(conn, tmp_path):
    path = tmp_path / "dividend.csv"
    csv_file(
        path,
        accounting_import.ZERODHA_DIVIDEND_HEADERS,
        [["ACME", "2026-01-03", "2", "1.5", "3"]],
    )
    parsed = accounting_import.parse_zerodha_dividends(path)
    accounting_import.store(conn, "zerodha", parsed)
    overlap = tmp_path / "overlap.csv"
    csv_file(
        overlap,
        accounting_import.ZERODHA_DIVIDEND_HEADERS,
        [["ACME", "2026-01-03", "2", "1.5", "3"], ["OTHER", "2026-02-01", 1, 2, 2]],
    )
    accounting_import.store(conn, "zerodha", accounting_import.parse_zerodha_dividends(overlap))
    assert conn.execute("SELECT count(*) FROM portfolio_income").fetchone()[0] == 2
    assert conn.execute("SELECT DISTINCT date_evidence FROM portfolio_income").fetchall() == [
        ("SUBSTITUTED_EX_DATE",)
    ]


def vested_workbook(path: Path):
    wb = Workbook()
    all_tx = wb.active
    all_tx.title = "All Transactions"
    all_tx.append(accounting_import.VESTED_SHEETS["All Transactions"])
    all_tx.append(["2026-01-01", "09:00:00", "CSR", 100, 100, "deposit"])
    all_tx.append(["2026-01-02", "09:00:00", "SPUR", 20, 80, "buy"])
    all_tx.append(["2026-01-03", "09:00:00", "DIV", 1, 81, "dividend"])
    all_tx.append(["2026-01-03", "09:00:01", "DIVTAX", -0.2, 80.8, "tax"])
    all_tx.append(["2026-01-04", "09:00:00", "FEE", 0.5, 80.3, "fee"])
    trades = wb.create_sheet("Trades")
    trades.append(accounting_import.VESTED_SHEETS["Trades"])
    trades.append(["2026-01-02", "09:00:00", "Acme", "ACME", "Buy", "Market", 2, 10, 20, 0])
    transfers = wb.create_sheet("Transfers")
    transfers.append(accounting_import.VESTED_SHEETS["Transfers"])
    transfers.append(["2026-01-01", "09:00:00", "Deposit", 100])
    income = wb.create_sheet("Income")
    income.append(accounting_import.VESTED_SHEETS["Income"])
    income.append(["2026-01-03", "09:00:00", "Dividend", "ACME", 1])
    income.append(["2026-01-03", "09:00:01", "Tax", "ACME", -0.2])
    wb.save(path)


def test_vested_import_splits_events_and_is_atomic(conn, tmp_path):
    path = tmp_path / "vested.xlsx"
    vested_workbook(path)
    parsed = accounting_import.parse_vested_transactions(path)
    result = accounting_import.store(conn, "vested", parsed)
    assert result["status"] == "stored"
    assert conn.execute("SELECT count(*) FROM portfolio_transaction").fetchone()[0] == 1
    assert conn.execute("SELECT count(*) FROM portfolio_cash_flow").fetchone()[0] == 1
    assert conn.execute("SELECT count(*) FROM portfolio_income").fetchone()[0] == 1
    assert conn.execute("SELECT count(*) FROM portfolio_tax").fetchone()[0] == 1
    assert conn.execute("SELECT count(*) FROM portfolio_fee").fetchone()[0] == 1

    bad = accounting_import.parse_vested_transactions(path)
    bad["fees"][0]["amount"] = -1
    bad["content_sha256"] = "f" * 64
    with pytest.raises(duckdb.ConstraintException):
        accounting_import.store(conn, "vested", bad)
    assert conn.execute("SELECT count(*) FROM accounting_import_run").fetchone()[0] == 1


def test_vested_cross_sheet_count_drift_fails_closed(tmp_path):
    path = tmp_path / "vested-drift.xlsx"
    vested_workbook(path)
    workbook = __import__("openpyxl").load_workbook(path)
    workbook["Transfers"].delete_rows(2)
    workbook.save(path)
    with pytest.raises(accounting_import.AccountingImportError, match="event counts"):
        accounting_import.parse_vested_transactions(path)


def test_vested_holdings_create_source_valuations(conn, tmp_path):
    path = tmp_path / "holdings.xlsx"
    workbook = Workbook()
    user = workbook.active
    user.title = "User Details"
    user.append(["Period", "User", "Govt ID", "DW Account Number", "Email"])
    user.append(["As of 28 Aug 2026", "private", "private", "private", "private"])
    summary = workbook.create_sheet("Summary")
    summary.append(
        [
            "Current Equity Value (USD)",
            "Total Amount Invested (USD)",
            "Investment Returns (USD)",
            "Investment Returns (%)",
        ]
    )
    summary.append([12, 10, 2, 20])
    holdings = workbook.create_sheet("Holdings")
    from invest import vested

    holdings.append(vested.HEADERS)
    holdings.append(["Acme", "ACME", 2, 6, 12, 5, 10, 2, 20, 0, 0])
    workbook.save(path)
    parsed = accounting_import.parse_vested_holdings(path)
    accounting_import.store(conn, "vested", parsed)
    assert conn.execute("SELECT value,valuation_evidence FROM portfolio_valuation").fetchone() == (
        pytest.approx(12),
        "SOURCE_SNAPSHOT",
    )


def test_schema_drift_and_provider_mismatch_fail_closed(conn, tmp_path):
    path = tmp_path / "trade.csv"
    csv_file(path, [*accounting_import.ZERODHA_TRADE_HEADERS, "new_column"], [["x"] * 14])
    with pytest.raises(accounting_import.AccountingImportError, match="headers"):
        accounting_import.parse_zerodha_tradebook(path)

    good = tmp_path / "dividend.csv"
    csv_file(good, accounting_import.ZERODHA_DIVIDEND_HEADERS, [["ACME", "2026-01-01", 1, 1, 1]])
    with pytest.raises(accounting_import.AccountingImportError, match="provider"):
        accounting_import.store(conn, "vested", accounting_import.parse_zerodha_dividends(good))
