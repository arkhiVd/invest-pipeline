from datetime import UTC, datetime

import duckdb
import pytest
from openpyxl import Workbook

from invest import db, vested


def workbook(path, *, ticker="ACME", summary=12.0):
    wb = Workbook()
    user = wb.active
    user.title = "User Details"
    user.append(["Period", "User", "Govt ID", "DW Account Number", "Email"])
    user.append(["As of 28 Aug 2026", "Private", "secret", "secret", "synthetic@example.invalid"])
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.append(
        [
            "Current Equity Value (USD)",
            "Total Amount Invested (USD)",
            "Investment Returns (USD)",
            "Investment Returns (%)",
        ]
    )
    summary_sheet.append([summary, 10.0, 2.0, 20.0])
    holdings = wb.create_sheet("Holdings")
    holdings.append(vested.HEADERS)
    holdings.append(["Acme Inc", ticker, 2.0, 6.0, 12.0, 5.0, 10.0, 2.0, 20.0, 0.0, 0.0])
    wb.save(path)


def test_store_replay_integrity_and_privacy(tmp_path):
    path = tmp_path / "holdings.xlsx"
    workbook(path)
    conn = duckdb.connect()
    db.init_schema(conn)
    first = vested.store(conn, path, now=datetime(2026, 8, 28, tzinfo=UTC))
    second = vested.store(conn, path, now=datetime(2026, 8, 29, tzinfo=UTC))
    assert first["status"] == "stored"
    assert second["status"] == "duplicate"
    assert vested.integrity(conn, first["run_id"])
    assert conn.execute("select count(*) from vested_holding").fetchone()[0] == 1
    columns = {
        row[0].lower()
        for row in conn.execute(
            "select column_name from information_schema.columns where table_name like 'vested%'"
        ).fetchall()
    }
    assert not columns & {"user", "email", "govt_id", "account_number"}
    conn.close()


def test_parser_fails_closed_on_drift_and_bad_totals(tmp_path):
    bad_total = tmp_path / "bad-total.xlsx"
    workbook(bad_total, summary=99.0)
    with pytest.raises(vested.VestedError, match="reconcile"):
        vested.parse(bad_total)
    bad_ticker = tmp_path / "bad-ticker.xlsx"
    workbook(bad_ticker, ticker="bad ticker")
    with pytest.raises(vested.VestedError, match="identity"):
        vested.parse(bad_ticker)
