"""T1.6/T1.6b acceptance: export smoke test against a seeded synthetic DB."""

import re
from datetime import date
from datetime import datetime as dt

import duckdb
from openpyxl import load_workbook

from invest import db, export, metrics

END = date(2025, 8, 31)
CALC_AT = dt(2025, 8, 31, 12, 0)


def lfr(sy, sm, base, rets):
    y, m, lvl, out = sy, sm, base, {}
    for r in rets:
        out[(y, m)] = lvl
        lvl *= 1 + r
        m += 1
        if m > 12:
            y, m = y + 1, 1
    out[(y, m)] = lvl
    return out


def pairs(levels):
    return [(metrics.month_end_date(k), v) for k, v in sorted(levels.items())]


BENCH = [x for _ in range(6) for x in [0.03, -0.01, 0.02, 0.04, -0.02, 0.01]]


def seeded_db():
    conn = duckdb.connect(":memory:")
    db.init_schema(conn)
    db.upsert_scheme(conn, scheme_code=900002, name="bench")
    db.upsert_navs(conn, 900002, pairs(lfr(2022, 8, 100.0, BENCH)))
    # full-history aggressive fund (beta 1.2)
    db.upsert_scheme(
        conn,
        scheme_code=201,
        name="A",
        display_name="Fund Alpha",
        category="Equity Scheme - Mid Cap Fund",
    )
    db.upsert_navs(conn, 201, pairs(lfr(2022, 8, 50.0, [1.2 * b for b in BENCH])))
    # ~14-month fund -> no 3Y but yes 1Y (needs >= 11 monthly returns)
    db.upsert_scheme(
        conn,
        scheme_code=202,
        name="B",
        display_name="Fund Beta",
        category="Equity Scheme - Mid Cap Fund",
    )
    db.upsert_navs(conn, 202, pairs(lfr(2024, 6, 10.0, [0.01] * 13)))
    # brand-new fund -> fully blank row
    db.upsert_scheme(
        conn,
        scheme_code=203,
        name="C",
        display_name="Fund Gamma",
        category="Equity Scheme - Mid Cap Fund",
    )
    db.upsert_navs(conn, 203, pairs(lfr(2025, 7, 10.0, [0.01] * 1)))
    metrics.run(conn, calculated_at=CALC_AT)
    return conn


def test_export_smoke(tmp_path):
    conn = seeded_db()
    out = tmp_path / "MFs_export.xlsx"
    summary = export.export_workbook(conn, out, END)
    assert summary["rows"] == 3 and summary["computed_3y"] == 1
    assert summary["computed_1y"] == 1

    wb = load_workbook(out)  # raises if the file does not open clean
    assert wb.sheetnames == ["Sheet 1"]
    ws = wb["Sheet 1"]
    assert [ws.cell(row=1, column=i).value for i in range(1, 20)] == export.HEADERS
    names = {ws.cell(row=r, column=1).value: r for r in range(2, 5)}
    r_a, r_b, r_g = names["Fund Alpha"], names["Fund Beta"], names["Fund Gamma"]

    # percent-point convention: returns/SD stored as e.g. 21.31 not 0.2131
    v = ws.cell(row=r_a, column=4).value
    assert isinstance(v, float) and 1 < v < 100
    # window column: Alpha=3Y, Beta=1Y fallback, Gamma blank
    assert ws.cell(row=r_a, column=18).value == "3Y"
    assert ws.cell(row=r_b, column=18).value == "1Y"
    assert ws.cell(row=r_g, column=18).value is None
    assert ws.cell(row=r_g, column=4).value is None
    # Beta has 1Y metrics filled where Alpha's row uses its 3Y values
    assert ws.cell(row=r_b, column=4).value is not None

    # conditional formatting replicas present (6 groups incl. text + sharpe)
    kinds = sorted(
        (str(cf.sqref), rule.type) for cf in ws.conditional_formatting for rule in cf.rules
    )
    assert ("A1:XFD1048576", "containsText") in kinds
    assert sum(1 for _, t in kinds if t == "colorScale") >= 6
    assert ("S2:S1048576", "cellIs") in kinds


def test_sorting_styling_and_filter(tmp_path):
    conn = seeded_db()
    out = tmp_path / "MFs_export.xlsx"
    export.export_workbook(conn, out, END)
    ws = load_workbook(out)["Sheet 1"]

    # category blocks A→Z, best return first inside block, no-data funds last:
    # Alpha (~19% p.a.) > Beta (1Y only, ~12%); Gamma has no data -> last
    order = [ws.cell(row=r, column=1).value for r in range(2, 5)]
    assert order == ["Fund Alpha", "Fund Beta", "Fund Gamma"]

    # autofilter over the full table for interactive sort/filter
    assert ws.auto_filter.ref == f"A1:S{ws.max_row}"
    # header styling: dark fill + white bold text
    h = ws.cell(row=1, column=4)
    assert h.fill.fgColor.rgb == "FF305496" and h.font.color.rgb == "FFFFFFFF"
    # borders on data cells; window chips colored per lookback
    assert ws.cell(row=2, column=3).border.left.style == "thin"
    assert ws.cell(row=2, column=18).fill.fgColor.rgb == "FFC6EFCE"
    assert ws.cell(row=3, column=18).fill.fgColor.rgb == "FFFFEB9C"
    # no-data row muted gray italic
    g_font = ws.cell(row=4, column=1).font
    assert g_font.italic and g_font.color.rgb == "FF9C9C9C"


def test_diff_string_mfanalysis2_convention():
    f = export._diff_string
    assert f("Outperformer", 0.1157, 0.1120) == "0.4 (O)"
    assert f("Underperformer", 0.064, 0.241) == "-17.7 (U)"
    assert f("At Category", 0.1, 0.0999) == "0.0"
    assert f(None, 0.1, 0.1) is None
    assert re.fullmatch(r"-?\d+\.\d+ \([OU]\)", f("Outperformer", 0.25, 0.20))


def test_category_label_mapping():
    f = export.category_label
    assert f("Equity Scheme - Mid Cap Fund") == "Mid Cap"
    assert f("Debt Scheme - Ultra Short Duration Fund") == "Ultra Short Duration"
    assert f(None) is None
